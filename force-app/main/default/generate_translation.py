#!/usr/bin/env python3
"""
=============================================================================
  Salesforce Translation Generator (v2)
  
  Generates the correct metadata files for Salesforce translations:
  - CustomObjectTranslation files for fields, picklists, objects
    → objectTranslations/<Object>-<lang>/<Object>-<lang>.objectTranslation-meta.xml
    → objectTranslations/<Object>-<lang>/<Field>.fieldTranslation-meta.xml
  
=============================================================================

Usage:
    python3 generate_translation.py --input <excel_file> --lang <language_code>
    python3 generate_translation.py --input translations.xlsx --lang fr --output ./force-app/main/default/

Excel Format (5 columns):
    Type             | ObjectOrEntity           | ApiName              | translation         | ANGLAIS
    CustomObject     | Operational_Task__c      | Operational_Task__c  | Mon Objet           | My Object
    CustomField      | Case                     | Complaint_reason__c  | Raison reclamation  | Complaint Reason
    PicklistValue    | Case.Complaint_reason__c | Agreement Breach     | Non respect accord  | Agreement Breach
    CustomFieldHelp  | Case                     | Complaint_reason__c  | Texte d'aide...     | Help text...
"""

import argparse
import sys
import os


# =============================================================================
#  STEP 1 - READ & CLEAN EXCEL
# =============================================================================

def read_excel(filepath):
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("ERROR: Excel file is empty.")
        sys.exit(1)

    data = rows[1:]
    processed = []
    skipped = []

    type_corrections = {
        "customobeject": "CustomObject",
        "customobject": "CustomObject",
        "customfield": "CustomField",
        "customfieldhelp": "CustomFieldHelp",
        "picklistvalue": "PicklistValue",
    }

    for i, row in enumerate(data, start=2):
        row = list(row) + [None] * (5 - len(row))
        raw_type, obj_entity, api_name, translation, english = row[:5]

        if all(v is None for v in row[:5]):
            continue

        if not raw_type or str(raw_type).strip() == "":
            skipped.append({
                "row": i,
                "reason": "Type manquant",
                "english": str(english or "").strip(),
                "translation": str(translation or "").strip(),
            })
            continue

        if not translation or str(translation).strip() == "":
            skipped.append({
                "row": i,
                "reason": "Traduction manquante",
                "english": str(english or "").strip(),
            })
            continue

        clean_type = str(raw_type).strip().replace("\t", "")
        clean_obj = str(obj_entity).strip().replace("\t", "") if obj_entity else ""
        clean_api = str(api_name).strip().replace("\t", "") if api_name else ""
        clean_translation = str(translation).strip()
        clean_english = str(english).strip() if english else ""

        normalized = clean_type.lower().replace(" ", "").replace("_", "")
        if normalized in type_corrections:
            clean_type = type_corrections[normalized]
        elif clean_type not in ("CustomObject", "CustomField", "CustomFieldHelp", "PicklistValue"):
            skipped.append({
                "row": i,
                "reason": f"Type inconnu: '{clean_type}'",
                "english": clean_english,
            })
            continue

        if not clean_obj:
            skipped.append({"row": i, "reason": "ObjectOrEntity manquant", "english": clean_english})
            continue

        if not clean_api:
            skipped.append({"row": i, "reason": "ApiName manquant", "english": clean_english})
            continue

        processed.append({
            "row": i,
            "type": clean_type,
            "object_or_entity": clean_obj,
            "api_name": clean_api,
            "translation": clean_translation,
            "english": clean_english,
        })

    return processed, skipped


# =============================================================================
#  STEP 2 - GROUP BY OBJECT
# =============================================================================

def group_by_object(rows):
    """
    Groups all rows by their parent object.
    Returns a dict: { "Case": { "fields": [...], "picklists": {...}, "object_label": ... }, ... }
    """
    objects = {}

    for row in rows:
        t = row["type"]
        obj = row["object_or_entity"]
        api = row["api_name"]

        if t == "CustomObject":
            # The object itself
            obj_name = row["api_name"]
            if obj_name not in objects:
                objects[obj_name] = {"fields": {}, "object_label": None}
            objects[obj_name]["object_label"] = row["translation"]

        elif t == "CustomField":
            if obj not in objects:
                objects[obj] = {"fields": {}, "object_label": None}
            if api not in objects[obj]["fields"]:
                objects[obj]["fields"][api] = {"label": None, "help": None, "picklist_values": []}
            objects[obj]["fields"][api]["label"] = row["translation"]

        elif t == "CustomFieldHelp":
            # obj might be "Case.Complaint_justified__c" or just "Case"
            # We need just the parent object
            if "." in obj:
                obj = obj.split(".")[0]
            if obj not in objects:
                objects[obj] = {"fields": {}, "object_label": None}
            if api not in objects[obj]["fields"]:
                objects[obj]["fields"][api] = {"label": None, "help": None, "picklist_values": []}
            objects[obj]["fields"][api]["help"] = row["translation"]

        elif t == "PicklistValue":
            # obj is like "Case.Complaint_reason__c"
            parts = obj.split(".", 1)
            if len(parts) == 2:
                parent_obj, field_api = parts
            else:
                parent_obj = obj
                field_api = api  # fallback
                continue  # skip malformed

            if parent_obj not in objects:
                objects[parent_obj] = {"fields": {}, "object_label": None}
            if field_api not in objects[parent_obj]["fields"]:
                objects[parent_obj]["fields"][field_api] = {"label": None, "help": None, "picklist_values": []}
            objects[parent_obj]["fields"][field_api]["picklist_values"].append({
                "masterLabel": api,
                "translation": row["translation"],
            })

    return objects


# =============================================================================
#  STEP 3 - GENERATE FILES
# =============================================================================

def escape_xml(text):
    return (text
            .replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;"))


def generate_field_translation_xml(field_api, field_data):
    """
    Generates a .fieldTranslation-meta.xml file content.
    
    Format:
    <?xml version="1.0" encoding="UTF-8"?>
    <CustomFieldTranslation xmlns="http://soap.sforce.com/2006/04/metadata">
        <help>...</help>           (if exists)
        <label>...</label>         (if exists)
        <name>Field__c</name>
        <picklistValues>           (for each picklist value)
            <masterLabel>...</masterLabel>
            <translation>...</translation>
        </picklistValues>
    </CustomFieldTranslation>
    """
    lines = []
    lines.append('<?xml version="1.0" encoding="UTF-8"?>')
    lines.append('<CustomFieldTranslation xmlns="http://soap.sforce.com/2006/04/metadata">')

    if field_data.get("help"):
        lines.append(f"    <help>{escape_xml(field_data['help'])}</help>")

    if field_data.get("label"):
        lines.append(f"    <label>{escape_xml(field_data['label'])}</label>")

    lines.append(f"    <name>{escape_xml(field_api)}</name>")

    for pv in sorted(field_data.get("picklist_values", []), key=lambda x: x["masterLabel"]):
        lines.append("    <picklistValues>")
        lines.append(f"        <masterLabel>{escape_xml(pv['masterLabel'])}</masterLabel>")
        lines.append(f"        <translation>{escape_xml(pv['translation'])}</translation>")
        lines.append("    </picklistValues>")

    lines.append("</CustomFieldTranslation>")
    return "\n".join(lines)


def generate_object_translation_xml(obj_name, obj_data):
    """
    Generates the main .objectTranslation-meta.xml file.
    
    Format:
    <?xml version="1.0" encoding="UTF-8"?>
    <CustomObjectTranslation xmlns="http://soap.sforce.com/2006/04/metadata">
        <name>Object__c-fr</name>     (if custom object label exists)
        <label>...</label>
    </CustomObjectTranslation>
    """
    lines = []
    lines.append('<?xml version="1.0" encoding="UTF-8"?>')
    lines.append('<CustomObjectTranslation xmlns="http://soap.sforce.com/2006/04/metadata">')

    if obj_data.get("object_label"):
        lines.append(f"    <label>{escape_xml(obj_data['object_label'])}</label>")

    lines.append("</CustomObjectTranslation>")
    return "\n".join(lines)


def generate_all_files(objects, lang, output_dir):
    """
    Generates the correct SFDX source folder structure:
    
    objectTranslations/
        Case-fr/
            Case-fr.objectTranslation-meta.xml
            Complaint_reason__c.fieldTranslation-meta.xml
            Operational_Status__c.fieldTranslation-meta.xml
        Operational_Task__c-fr/
            Operational_Task__c-fr.objectTranslation-meta.xml
            Status__c.fieldTranslation-meta.xml
    """
    translations_dir = os.path.join(output_dir, "objectTranslations")
    files_created = []

    for obj_name, obj_data in sorted(objects.items()):
        # Create folder: objectTranslations/<Object>-<lang>/
        folder_name = f"{obj_name}-{lang}"
        folder_path = os.path.join(translations_dir, folder_name)
        os.makedirs(folder_path, exist_ok=True)

        # Generate main objectTranslation file
        obj_xml = generate_object_translation_xml(obj_name, obj_data)
        obj_file = os.path.join(folder_path, f"{folder_name}.objectTranslation-meta.xml")
        with open(obj_file, "w", encoding="utf-8") as f:
            f.write(obj_xml)
        files_created.append(obj_file)

        # Generate field translation files
        for field_api, field_data in sorted(obj_data["fields"].items()):
            field_xml = generate_field_translation_xml(field_api, field_data)
            field_file = os.path.join(folder_path, f"{field_api}.fieldTranslation-meta.xml")
            with open(field_file, "w", encoding="utf-8") as f:
                f.write(field_xml)
            files_created.append(field_file)

    return files_created


# =============================================================================
#  STEP 4 - REPORT
# =============================================================================

def print_report(processed, skipped, files_created, output_dir, lang):
    print("\n" + "=" * 60)
    print("  TRANSLATION GENERATION REPORT")
    print("=" * 60)

    type_counts = {}
    for row in processed:
        t = row["type"]
        type_counts[t] = type_counts.get(t, 0) + 1

    print(f"\n  Lines processed: {len(processed)}")
    for t, count in sorted(type_counts.items()):
        print(f"     - {t}: {count}")

    print(f"\n  Skipped: {len(skipped)} lines")
    if skipped:
        for item in skipped:
            english = item.get("english", "???")
            reason = item["reason"]
            print(f"     - Row {item['row']}: \"{english}\" -> {reason}")

    print(f"\n  Files created: {len(files_created)}")
    for f in files_created:
        print(f"     - {f}")

    print(f"\n  Output directory: {output_dir}")
    print("=" * 60)
    print()


# =============================================================================
#  MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Generate Salesforce translation files from Excel"
    )
    parser.add_argument("--input", "-i", required=True, help="Path to the Excel file")
    parser.add_argument("--lang", "-l", required=True, help="Language code (e.g., fr, es, de)")
    parser.add_argument("--output", "-o", default=".", help="Output directory (default: current directory)")

    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"ERROR: File not found: {args.input}")
        sys.exit(1)

    # Step 1 - Read & Clean
    print(f"\nReading: {args.input}")
    processed, skipped = read_excel(args.input)

    if not processed:
        print("ERROR: No valid rows found.")
        print_report(processed, skipped, [], args.output, args.lang)
        sys.exit(1)

    # Step 2 - Group by object
    objects = group_by_object(processed)

    # Step 3 - Generate files
    files_created = generate_all_files(objects, args.lang, args.output)

    # Step 4 - Report
    print_report(processed, skipped, files_created, args.output, args.lang)

    print("Done! Next steps:")
    print(f"  1. Copy the objectTranslations/ folder into your SFDX project under force-app/main/default/")
    print(f"  2. Deploy with: sf project deploy start --metadata CustomObjectTranslation --target-org mon-org")
    print()


if __name__ == "__main__":
    main()